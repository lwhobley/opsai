from dotenv import load_dotenv
from pathlib import Path
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

from fastapi import FastAPI, APIRouter, Depends, HTTPException, UploadFile, File, Request
from fastapi.responses import JSONResponse, StreamingResponse
from starlette.middleware.cors import CORSMiddleware
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_, desc
from sqlalchemy.orm import selectinload
import os
import logging
from pydantic import BaseModel, Field
from typing import List, Optional
from datetime import datetime, timezone, timedelta
import jwt
import bcrypt
import json
import pandas as pd
import pdfplumber
from io import BytesIO
from google import genai as genai_client
import httpx

from database import get_db, engine, Base, AsyncSessionLocal
from models import (
    User, InventoryItem, InventoryCount, KitchenInventoryItem, 
    KitchenInventoryCount, Purchase, Sale, MenuItem, MenuItemIngredient,
    ToastIntegration,
    WasteLog, PushSubscription
)
from pywebpush import webpush, WebPushException
from apscheduler.schedulers.asyncio import AsyncIOScheduler

# ── Simple in-memory login rate limiter ──────────────────────────────────────
from contextlib import asynccontextmanager
from collections import defaultdict
import time as _time

_login_attempts: dict = defaultdict(list)   # ip -> [timestamps]
LOGIN_RATE_LIMIT  = int(os.environ.get("LOGIN_RATE_LIMIT", "10"))   # max attempts
LOGIN_RATE_WINDOW = int(os.environ.get("LOGIN_RATE_WINDOW", "300"))  # per N seconds (5 min)

def _check_rate_limit(ip: str) -> None:
    now = _time.monotonic()
    window_start = now - LOGIN_RATE_WINDOW
    _login_attempts[ip] = [t for t in _login_attempts[ip] if t > window_start]
    if len(_login_attempts[ip]) >= LOGIN_RATE_LIMIT:
        raise HTTPException(status_code=429, detail="Too many login attempts. Please wait 5 minutes.")
    _login_attempts[ip].append(now)

def _clear_rate_limit(ip: str) -> None:
    _login_attempts.pop(ip, None)

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Configure Gemini
_GEMINI_API_KEY = os.environ.get('GEMINI_API_KEY', '')

# JWT Config
JWT_SECRET = os.environ.get('JWT_SECRET')
if not JWT_SECRET:
    raise RuntimeError("JWT_SECRET environment variable is required and must not be empty")
JWT_ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24  # 24 hours

@asynccontextmanager
async def lifespan(app: FastAPI):
    # ── Startup ──────────────────────────────────────────────
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)

    async with AsyncSessionLocal() as db:
        admin_pin = os.environ.get("ADMIN_PIN")
        if not admin_pin:
            raise RuntimeError("ADMIN_PIN environment variable is required to seed the admin user.")
        result = await db.execute(select(User).where(User.role == "admin"))
        admin = result.scalar_one_or_none()
        if not admin:
            admin_user = User(name="Admin", pin_hash=hash_pin(admin_pin), role="admin")
            db.add(admin_user)
            await db.commit()
            logger.info("Admin user seeded from ADMIN_PIN env var.")
        else:
            # Always sync admin PIN from env var so deploys fix stale hashes
            admin.pin_hash = hash_pin(admin_pin)
            await db.commit()
            logger.info("Admin PIN synced from ADMIN_PIN env var.")

    # ── Scheduler: low-stock push alerts ─────────────────────
    scheduler = AsyncIOScheduler()
    scheduler.add_job(
        _scheduled_low_stock_check,
        trigger="cron",
        hour=7, minute=0,          # 7 AM daily (UTC)
        id="low_stock_check",
        replace_existing=True,
    )
    scheduler.start()
    logger.info("APScheduler started — low-stock check runs daily at 07:00 UTC")

    yield  # app is running

    # ── Shutdown ─────────────────────────────────────────────
    scheduler.shutdown(wait=False)
    await engine.dispose()

app = FastAPI(title="Ops AI - Restaurant Operations", lifespan=lifespan)
api_router = APIRouter(prefix="/api")

# Pydantic Models
class PinLogin(BaseModel):
    pin: str

class UserCreate(BaseModel):
    name: str
    pin: str
    role: str = "staff"

class UserResponse(BaseModel):
    id: str
    name: str
    role: str
    is_active: bool

class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: UserResponse

class InventoryItemCreate(BaseModel):
    name: str
    category: Optional[str] = None
    subcategory: Optional[str] = None
    location: Optional[str] = None
    section: Optional[str] = None
    bottle_size_ml: Optional[int] = None
    cost_per_unit: float = 0.0
    display_order: int = 0
    count_priority: int = 0

class InventoryItemResponse(BaseModel):
    id: str
    name: str
    category: Optional[str]
    subcategory: Optional[str]
    location: Optional[str]
    section: Optional[str]
    bottle_size_ml: Optional[int]
    cost_per_unit: float
    display_order: int
    count_priority: int
    is_active: bool
    latest_count: Optional[int] = None

class InventoryCountCreate(BaseModel):
    item_id: str
    level_percentage: int

class KitchenItemCreate(BaseModel):
    name: str
    unit: Optional[str] = None
    location: Optional[str] = None
    station: Optional[str] = None
    cost_per_unit: float = 0.0
    vendor: Optional[str] = None
    display_order: int = 0
    par_level: float = 0.0

class KitchenItemResponse(BaseModel):
    id: str
    name: str
    unit: Optional[str]
    location: Optional[str]
    station: Optional[str]
    cost_per_unit: float
    vendor: Optional[str]
    display_order: int
    par_level: float
    is_active: bool
    latest_count: Optional[float] = None

class KitchenCountCreate(BaseModel):
    item_id: str
    quantity: float

class PurchaseCreate(BaseModel):
    item_name: str
    purchase_type: Optional[str] = None   # canonical field: bar | kitchen | supply | other
    quantity: float
    total_cost: float
    date: Optional[datetime] = None
    item_type: Optional[str] = None       # legacy alias — resolved server-side

class SaleCreate(BaseModel):
    date: datetime
    total_sales: float
    bar_sales: float = 0.0
    food_sales: float = 0.0

class MenuItemCreate(BaseModel):
    name: str
    category: Optional[str] = None
    price: float

class MenuIngredientCreate(BaseModel):
    menu_item_id: str
    ingredient_name: str
    quantity_used: float
    unit: Optional[str] = None
    cost_per_unit: float = 0.0

class AIInsightRequest(BaseModel):
    include_bar: bool = True
    include_kitchen: bool = True
    date_range_days: int = 7

# Auth helpers
def hash_pin(pin: str) -> str:
    return bcrypt.hashpw(pin.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_pin(pin: str, hashed: str) -> bool:
    return bcrypt.checkpw(pin.encode('utf-8'), hashed.encode('utf-8'))

def create_access_token(user_id: str, role: str) -> str:
    payload = {
        "sub": user_id,
        "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES),
        "type": "access"
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(request: Request, db: AsyncSession = Depends(get_db)) -> User:
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Invalid token type")
        result = await db.execute(select(User).where(User.id == payload["sub"]))
        user = result.scalar_one_or_none()
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

async def require_admin(user: User = Depends(get_current_user)) -> User:
    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return user

async def require_manager(user: User = Depends(get_current_user)) -> User:
    if user.role not in ["admin", "manager"]:
        raise HTTPException(status_code=403, detail="Manager access required")
    return user

# Auth endpoints
@api_router.post("/auth/login", response_model=TokenResponse)
async def login(data: PinLogin, request: Request, db: AsyncSession = Depends(get_db)):
    client_ip = request.client.host if request.client else "unknown"
    _check_rate_limit(client_ip)

    result = await db.execute(select(User).where(User.is_active == True))
    users = result.scalars().all()
    
    for user in users:
        if verify_pin(data.pin, user.pin_hash):
            _clear_rate_limit(client_ip)   # reset on success
            token = create_access_token(user.id, user.role)
            response = JSONResponse(content={
                "access_token": token,
                "token_type": "bearer",
                "user": {
                    "id": user.id,
                    "name": user.name,
                    "role": user.role,
                    "is_active": user.is_active
                }
            })
            response.set_cookie(
                key="access_token", value=token, httponly=True,
                secure=os.environ.get("COOKIE_SECURE", "true").lower() == "true",
                samesite="lax", max_age=86400, path="/"
            )
            return response
    
    raise HTTPException(status_code=401, detail="Invalid PIN")

@api_router.post("/auth/logout")
async def logout():
    response = JSONResponse(content={"message": "Logged out"})
    response.delete_cookie("access_token")
    return response

@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(user: User = Depends(get_current_user)):
    return UserResponse(id=user.id, name=user.name, role=user.role, is_active=user.is_active)

# User management
@api_router.post("/users", response_model=UserResponse)
async def create_user(data: UserCreate, db: AsyncSession = Depends(get_db), admin: User = Depends(require_admin)):
    user = User(name=data.name, pin_hash=hash_pin(data.pin), role=data.role)
    db.add(user)
    await db.commit()
    await db.refresh(user)
    return UserResponse(id=user.id, name=user.name, role=user.role, is_active=user.is_active)

@api_router.get("/users", response_model=List[UserResponse])
async def get_users(db: AsyncSession = Depends(get_db), admin: User = Depends(require_admin)):
    result = await db.execute(select(User).where(User.is_active == True))
    users = result.scalars().all()
    return [UserResponse(id=u.id, name=u.name, role=u.role, is_active=u.is_active) for u in users]

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, db: AsyncSession = Depends(get_db), admin: User = Depends(require_admin)):
    result = await db.execute(select(User).where(User.id == user_id))
    user = result.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    user.is_active = False
    await db.commit()
    return {"message": "User deactivated"}

class PinReset(BaseModel):
    pin: str

@api_router.put("/users/{user_id}/pin")
async def reset_user_pin(user_id: str, data: PinReset, db: AsyncSession = Depends(get_db), admin: User = Depends(require_admin)):
    if not data.pin or not data.pin.isdigit() or len(data.pin) not in (4, 6):
        raise HTTPException(status_code=400, detail="PIN must be exactly 4 or 6 digits")
    result = await db.execute(select(User).where(User.id == user_id, User.is_active == True))
    user = result.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    user.pin_hash = hash_pin(data.pin)
    await db.commit()
    return {"message": "PIN updated successfully"}

# Bar Inventory endpoints
@api_router.post("/inventory/bar/items", response_model=InventoryItemResponse)
async def create_bar_item(data: InventoryItemCreate, db: AsyncSession = Depends(get_db), user: User = Depends(require_manager)):
    item = InventoryItem(**data.model_dump())
    db.add(item)
    await db.commit()
    await db.refresh(item)
    return InventoryItemResponse(**{**item.__dict__, 'latest_count': None})

@api_router.get("/inventory/bar/items", response_model=List[InventoryItemResponse])
async def get_bar_items(location: Optional[str] = None, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    # Latest count per item via MAX(timestamp) subquery — eliminates N+1
    latest_ts_sub = (
        select(InventoryCount.item_id, func.max(InventoryCount.timestamp).label("max_ts"))
        .group_by(InventoryCount.item_id)
        .subquery()
    )
    latest_count_sub = (
        select(InventoryCount.item_id, InventoryCount.level_percentage)
        .join(latest_ts_sub, and_(
            InventoryCount.item_id == latest_ts_sub.c.item_id,
            InventoryCount.timestamp == latest_ts_sub.c.max_ts,
        ))
        .subquery()
    )
    query = (
        select(InventoryItem, latest_count_sub.c.level_percentage)
        .outerjoin(latest_count_sub, InventoryItem.id == latest_count_sub.c.item_id)
        .where(InventoryItem.is_active == True)
    )
    if location:
        query = query.where(InventoryItem.location == location)
    query = query.order_by(InventoryItem.location, InventoryItem.section, InventoryItem.display_order)
    result = await db.execute(query)
    rows = result.all()
    return [
        InventoryItemResponse(
            id=item.id, name=item.name, category=item.category, subcategory=item.subcategory,
            location=item.location, section=item.section, bottle_size_ml=item.bottle_size_ml,
            cost_per_unit=item.cost_per_unit, display_order=item.display_order,
            count_priority=item.count_priority, is_active=item.is_active, latest_count=latest_pct
        )
        for item, latest_pct in rows
    ]

@api_router.post("/inventory/bar/counts")
async def record_bar_count(data: InventoryCountCreate, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    count = InventoryCount(item_id=data.item_id, level_percentage=data.level_percentage, user_id=user.id)
    db.add(count)
    await db.commit()
    return {"message": "Count recorded", "id": count.id}

@api_router.post("/inventory/bar/counts/bulk")
async def record_bar_counts_bulk(counts: List[InventoryCountCreate], db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    for data in counts:
        count = InventoryCount(item_id=data.item_id, level_percentage=data.level_percentage, user_id=user.id)
        db.add(count)
    await db.commit()
    return {"message": f"{len(counts)} counts recorded"}

@api_router.delete("/inventory/bar/items/{item_id}")
async def delete_bar_item(item_id: str, db: AsyncSession = Depends(get_db), user: User = Depends(require_manager)):
    result = await db.execute(select(InventoryItem).where(InventoryItem.id == item_id))
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    item.is_active = False
    await db.commit()
    return {"message": "Item deactivated"}

# Kitchen Inventory endpoints
@api_router.post("/inventory/kitchen/items", response_model=KitchenItemResponse)
async def create_kitchen_item(data: KitchenItemCreate, db: AsyncSession = Depends(get_db), user: User = Depends(require_manager)):
    item = KitchenInventoryItem(**data.model_dump())
    db.add(item)
    await db.commit()
    await db.refresh(item)
    return KitchenItemResponse(**{**item.__dict__, 'latest_count': None})

@api_router.get("/inventory/kitchen/items", response_model=List[KitchenItemResponse])
async def get_kitchen_items(location: Optional[str] = None, station: Optional[str] = None, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    # Latest count per item via MAX(timestamp) subquery — eliminates N+1
    latest_ts_sub = (
        select(KitchenInventoryCount.item_id, func.max(KitchenInventoryCount.timestamp).label("max_ts"))
        .group_by(KitchenInventoryCount.item_id)
        .subquery()
    )
    latest_count_sub = (
        select(KitchenInventoryCount.item_id, KitchenInventoryCount.quantity)
        .join(latest_ts_sub, and_(
            KitchenInventoryCount.item_id == latest_ts_sub.c.item_id,
            KitchenInventoryCount.timestamp == latest_ts_sub.c.max_ts,
        ))
        .subquery()
    )
    query = (
        select(KitchenInventoryItem, latest_count_sub.c.quantity)
        .outerjoin(latest_count_sub, KitchenInventoryItem.id == latest_count_sub.c.item_id)
        .where(KitchenInventoryItem.is_active == True)
    )
    if location:
        query = query.where(KitchenInventoryItem.location == location)
    if station:
        query = query.where(KitchenInventoryItem.station == station)
    query = query.order_by(KitchenInventoryItem.location, KitchenInventoryItem.station, KitchenInventoryItem.display_order)
    result = await db.execute(query)
    rows = result.all()
    return [
        KitchenItemResponse(
            id=item.id, name=item.name, unit=item.unit, location=item.location,
            station=item.station, cost_per_unit=item.cost_per_unit, vendor=item.vendor,
            display_order=item.display_order, par_level=item.par_level, is_active=item.is_active,
            latest_count=latest_qty
        )
        for item, latest_qty in rows
    ]

@api_router.post("/inventory/kitchen/counts")
async def record_kitchen_count(data: KitchenCountCreate, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    count = KitchenInventoryCount(item_id=data.item_id, quantity=data.quantity, user_id=user.id)
    db.add(count)
    await db.commit()
    return {"message": "Count recorded", "id": count.id}

@api_router.post("/inventory/kitchen/counts/bulk")
async def record_kitchen_counts_bulk(counts: List[KitchenCountCreate], db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    for data in counts:
        count = KitchenInventoryCount(item_id=data.item_id, quantity=data.quantity, user_id=user.id)
        db.add(count)
    await db.commit()
    return {"message": f"{len(counts)} counts recorded"}

@api_router.delete("/inventory/kitchen/items/{item_id}")
async def delete_kitchen_item(item_id: str, db: AsyncSession = Depends(get_db), user: User = Depends(require_manager)):
    result = await db.execute(select(KitchenInventoryItem).where(KitchenInventoryItem.id == item_id))
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    item.is_active = False
    await db.commit()
    return {"message": "Item deactivated"}

# Purchases
@api_router.post("/purchases")
async def create_purchase(data: PurchaseCreate, db: AsyncSession = Depends(get_db), user: User = Depends(require_manager)):
    dump = data.model_dump()
    # Resolve canonical type — purchase_type wins, item_type is legacy alias
    resolved = dump.get('purchase_type') or dump.get('item_type') or 'other'
    dump['purchase_type'] = resolved
    dump['item_type'] = resolved          # keep DB column in sync
    purchase = Purchase(**dump)
    db.add(purchase)
    await db.commit()
    return {"message": "Purchase recorded", "id": purchase.id}

@api_router.get("/purchases")
async def get_purchases(days: int = 30, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    start_date = datetime.now(timezone.utc) - timedelta(days=days)
    result = await db.execute(
        select(Purchase).where(Purchase.date >= start_date).order_by(desc(Purchase.date))
    )
    purchases = result.scalars().all()
    return [{"id": p.id, "item_name": p.item_name, "item_type": p.item_type, 
             "quantity": p.quantity, "total_cost": p.total_cost, "date": p.date.isoformat(),
             "purchase_type": p.purchase_type} for p in purchases]

# Sales
@api_router.post("/sales")
async def create_sale(data: SaleCreate, db: AsyncSession = Depends(get_db), user: User = Depends(require_manager)):
    # Upsert by calendar date — enforce one record per day
    day_start = data.date.replace(hour=0, minute=0, second=0, microsecond=0)
    day_end   = day_start + timedelta(days=1)
    # Find any existing records for this day (guard against legacy duplicates with first())
    existing_result = await db.execute(
        select(Sale).where(Sale.date >= day_start, Sale.date < day_end).order_by(Sale.date)
    )
    all_existing = existing_result.scalars().all()
    # Delete extra duplicates if any exist (legacy data cleanup)
    for dup in all_existing[1:]:
        await db.delete(dup)
    existing = all_existing[0] if all_existing else None
    if existing:
        existing.total_sales = data.total_sales
        existing.bar_sales   = data.bar_sales
        existing.food_sales  = data.food_sales
        existing.date        = data.date
        await db.commit()
        return {"message": "Sale updated", "id": existing.id}
    sale = Sale(**data.model_dump())
    db.add(sale)
    await db.commit()
    return {"message": "Sale recorded", "id": sale.id}

@api_router.get("/sales")
async def get_sales(days: int = 30, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    start_date = datetime.now(timezone.utc) - timedelta(days=days)
    result = await db.execute(
        select(Sale).where(Sale.date >= start_date).order_by(desc(Sale.date))
    )
    sales = result.scalars().all()
    return [{"id": s.id, "date": s.date.isoformat(), "total_sales": s.total_sales,
             "bar_sales": s.bar_sales, "food_sales": s.food_sales} for s in sales]

@api_router.delete("/purchases/{purchase_id}")
async def delete_purchase(purchase_id: str, db: AsyncSession = Depends(get_db), user: User = Depends(require_manager)):
    result = await db.execute(select(Purchase).where(Purchase.id == purchase_id))
    p = result.scalar_one_or_none()
    if not p:
        raise HTTPException(status_code=404, detail="Purchase not found")
    await db.delete(p)
    await db.commit()
    return {"message": "Deleted"}

@api_router.delete("/sales/{sale_id}")
async def delete_sale(sale_id: str, db: AsyncSession = Depends(get_db), user: User = Depends(require_manager)):
    result = await db.execute(select(Sale).where(Sale.id == sale_id))
    s = result.scalar_one_or_none()
    if not s:
        raise HTTPException(status_code=404, detail="Sale not found")
    await db.delete(s)
    await db.commit()
    return {"message": "Deleted"}

# Settings — cost targets
@api_router.get("/settings/targets")
async def get_targets(user: User = Depends(get_current_user)):
    return {
        "pour_cost_target": float(os.environ.get("TARGET_POUR_COST_PCT", "20.0")),
        "food_cost_target": float(os.environ.get("TARGET_FOOD_COST_PCT", "30.0")),
    }

# Menu Items
@api_router.post("/menu/items")
async def create_menu_item(data: MenuItemCreate, db: AsyncSession = Depends(get_db), user: User = Depends(require_manager)):
    item = MenuItem(**data.model_dump())
    db.add(item)
    await db.commit()
    return {"message": "Menu item created", "id": item.id}

@api_router.get("/menu/items")
async def get_menu_items(db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    result = await db.execute(
        select(MenuItem).where(MenuItem.is_active == True).options(selectinload(MenuItem.ingredients))
    )
    items = result.scalars().all()
    response = []
    for item in items:
        total_cost = sum(i.quantity_used * i.cost_per_unit for i in item.ingredients)
        food_cost_pct = (total_cost / item.price * 100) if item.price > 0 else 0
        response.append({
            "id": item.id, "name": item.name, "category": item.category,
            "price": item.price, "total_cost": round(total_cost, 2),
            "food_cost_pct": round(food_cost_pct, 1),
            "profit_margin": round(item.price - total_cost, 2),
            "ingredients": [{"name": i.ingredient_name, "quantity": i.quantity_used, 
                           "unit": i.unit, "cost": i.cost_per_unit} for i in item.ingredients]
        })
    return response

@api_router.post("/menu/ingredients")
async def add_menu_ingredient(data: MenuIngredientCreate, db: AsyncSession = Depends(get_db), user: User = Depends(require_manager)):
    ingredient = MenuItemIngredient(**data.model_dump())
    db.add(ingredient)
    await db.commit()
    return {"message": "Ingredient added", "id": ingredient.id}

# Dashboard / Analytics
async def _compute_dashboard(days: int, db: AsyncSession) -> dict:
    """Shared dashboard logic — called by route and AI insights."""
    from sqlalchemy import text
    start_date = datetime.now(timezone.utc) - timedelta(days=days)

    # Sales totals
    sales_result = await db.execute(
        select(func.sum(Sale.total_sales), func.sum(Sale.bar_sales), func.sum(Sale.food_sales))
        .where(Sale.date >= start_date)
    )
    sales_data = sales_result.first()
    total_sales  = sales_data[0] or 0
    bar_sales    = sales_data[1] or 0
    food_sales   = sales_data[2] or 0

    # COGS totals — single query grouped by type
    cogs_result = await db.execute(
        select(Purchase.purchase_type, func.sum(Purchase.total_cost))
        .where(Purchase.date >= start_date)
        .group_by(Purchase.purchase_type)
    )
    cogs_by_type = {row[0]: row[1] or 0 for row in cogs_result.all()}
    total_purchases   = sum(cogs_by_type.values())
    bar_purchases     = cogs_by_type.get('bar', 0)
    kitchen_purchases = cogs_by_type.get('kitchen', 0)

    pour_cost_pct  = (bar_purchases    / bar_sales    * 100) if bar_sales    > 0 else 0
    food_cost_pct  = (kitchen_purchases / food_sales   * 100) if food_sales  > 0 else 0
    total_cogs_pct = (total_purchases   / total_sales  * 100) if total_sales > 0 else 0

    target_pour = float(os.environ.get("TARGET_POUR_COST_PCT", "20.0"))
    target_food = float(os.environ.get("TARGET_FOOD_COST_PCT", "30.0"))

    # Low stock bar items — single query using subquery for latest count per item
    bar_low_sq = (
        select(
            InventoryCount.item_id,
            func.max(InventoryCount.timestamp).label("latest_ts")
        )
        .group_by(InventoryCount.item_id)
        .subquery()
    )
    bar_low_result = await db.execute(
        select(InventoryItem.name, InventoryItem.location, InventoryCount.level_percentage)
        .join(bar_low_sq, InventoryItem.id == bar_low_sq.c.item_id)
        .join(InventoryCount, and_(
            InventoryCount.item_id == bar_low_sq.c.item_id,
            InventoryCount.timestamp == bar_low_sq.c.latest_ts
        ))
        .where(InventoryItem.is_active == True)
        .where(InventoryCount.level_percentage <= 25)
    )
    low_bar_items = [
        {"name": r[0], "location": r[1], "level": r[2]}
        for r in bar_low_result.all()
    ]

    # Low stock kitchen items — single query using subquery
    kit_low_sq = (
        select(
            KitchenInventoryCount.item_id,
            func.max(KitchenInventoryCount.timestamp).label("latest_ts")
        )
        .group_by(KitchenInventoryCount.item_id)
        .subquery()
    )
    kit_low_result = await db.execute(
        select(
            KitchenInventoryItem.name, KitchenInventoryItem.location,
            KitchenInventoryItem.par_level, KitchenInventoryCount.quantity
        )
        .join(kit_low_sq, KitchenInventoryItem.id == kit_low_sq.c.item_id)
        .join(KitchenInventoryCount, and_(
            KitchenInventoryCount.item_id == kit_low_sq.c.item_id,
            KitchenInventoryCount.timestamp == kit_low_sq.c.latest_ts
        ))
        .where(KitchenInventoryItem.is_active == True)
        .where(KitchenInventoryItem.par_level > 0)
        .where(KitchenInventoryCount.quantity < KitchenInventoryItem.par_level)
    )
    low_kitchen_items = [
        {"name": r[0], "location": r[1], "par_level": r[2], "quantity": r[3]}
        for r in kit_low_result.all()
    ]

    return {
        "period_days": days,
        "total_sales": round(total_sales, 2),
        "bar_sales":   round(bar_sales, 2),
        "food_sales":  round(food_sales, 2),
        "total_cogs":  round(total_purchases, 2),
        "bar_cogs":    round(bar_purchases, 2),
        "food_cogs":   round(kitchen_purchases, 2),
        "pour_cost_pct":   round(pour_cost_pct, 1),
        "food_cost_pct":   round(food_cost_pct, 1),
        "total_cogs_pct":  round(total_cogs_pct, 1),
        "low_bar_items":     low_bar_items[:10],
        "low_kitchen_items": low_kitchen_items[:10],
        "variance": {
            "target_pour_cost": target_pour,
            "target_food_cost": target_food,
            "pour_cost_variance": round(pour_cost_pct - target_pour, 1),
            "food_cost_variance": round(food_cost_pct - target_food, 1),
        },
    }

@api_router.get("/dashboard")
async def get_dashboard(days: int = 7, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    return await _compute_dashboard(days, db)

# AI Insights
@api_router.post("/ai/insights")
async def get_ai_insights(request: AIInsightRequest, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    dashboard_data = await _compute_dashboard(request.date_range_days, db)
    
    # Get menu costing data
    menu_result = await db.execute(
        select(MenuItem).where(MenuItem.is_active == True).options(selectinload(MenuItem.ingredients))
    )
    menu_items = menu_result.scalars().all()
    low_margin_items = []
    for item in menu_items:
        total_cost = sum(i.quantity_used * i.cost_per_unit for i in item.ingredients)
        margin_pct = ((item.price - total_cost) / item.price * 100) if item.price > 0 else 0
        if margin_pct < 60:
            low_margin_items.append({"name": item.name, "margin_pct": round(margin_pct, 1), "cost": round(total_cost, 2), "price": item.price})
    
    # Build context for AI
    context = {
        "period": f"Last {request.date_range_days} days",
        "sales": {
            "total": dashboard_data["total_sales"],
            "bar": dashboard_data["bar_sales"],
            "food": dashboard_data["food_sales"]
        },
        "costs": {
            "pour_cost_pct": dashboard_data["pour_cost_pct"],
            "food_cost_pct": dashboard_data["food_cost_pct"],
            "total_cogs_pct": dashboard_data["total_cogs_pct"]
        },
        "variance": dashboard_data["variance"],
        "low_stock_bar": dashboard_data["low_bar_items"],
        "low_stock_kitchen": dashboard_data["low_kitchen_items"],
        "low_margin_menu_items": low_margin_items[:5]
    }
    
    prompt = f"""You are an experienced restaurant General Manager analyzing operational data.
    
Analyze this data and provide actionable insights:
{json.dumps(context, indent=2)}

Target benchmarks:
- Pour cost should be under 20%
- Food cost should be under 30%
- Total COGS should be under 35%

Respond in this exact JSON format:
{{
    "key_issues": ["issue 1", "issue 2", "issue 3"],
    "likely_causes": ["cause 1", "cause 2", "cause 3"],
    "recommendations": [
        {{"title": "Action 1", "description": "Specific steps to take", "priority": "high"}},
        {{"title": "Action 2", "description": "Specific steps to take", "priority": "medium"}},
        {{"title": "Action 3", "description": "Specific steps to take", "priority": "low"}}
    ],
    "summary": "One sentence executive summary"
}}

Be direct, operational, and profit-focused. Write like an experienced GM."""

    try:
        _client = genai_client.Client(api_key=_GEMINI_API_KEY)
        response = _client.models.generate_content(
            model=os.environ.get('GEMINI_INSIGHTS_MODEL', 'gemini-2.0-flash-lite'),
            contents=prompt,
        )
        response_text = response.text.strip()
        
        # Clean up response
        if response_text.startswith("```json"):
            response_text = response_text[7:]
        if response_text.startswith("```"):
            response_text = response_text[3:]
        if response_text.endswith("```"):
            response_text = response_text[:-3]
        
        insights = json.loads(response_text.strip())
        return {"insights": insights, "context": context}
    except Exception as e:
        logger.error(f"AI insights error: {e}")
        return {
            "insights": {
                "key_issues": ["Unable to generate AI insights at this time"],
                "likely_causes": ["Service temporarily unavailable"],
                "recommendations": [{"title": "Try Again", "description": "Please refresh and try again", "priority": "medium"}],
                "summary": "AI analysis temporarily unavailable"
            },
            "context": context
        }

# File Upload / Import
@api_router.post("/import/inventory")
async def import_inventory(file: UploadFile = File(...), inventory_type: str = "bar", db: AsyncSession = Depends(get_db), user: User = Depends(require_manager)):
    content = await file.read()
    filename = file.filename.lower()
    
    try:
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            df = pd.read_excel(BytesIO(content))
        elif filename.endswith('.csv'):
            df = pd.read_csv(BytesIO(content))
        elif filename.endswith('.pdf'):
            with pdfplumber.open(BytesIO(content)) as pdf:
                tables = []
                for page in pdf.pages:
                    page_tables = page.extract_tables()
                    if page_tables:
                        tables.extend(page_tables)
                if not tables:
                    raise HTTPException(status_code=400, detail="No tables found in PDF")
                df = pd.DataFrame(tables[0][1:], columns=tables[0][0])
        else:
            raise HTTPException(status_code=400, detail="Unsupported file type. Use Excel, CSV, or PDF")
        
        # Normalize column names
        df.columns = df.columns.str.lower().str.strip()
        
        imported_count = 0
        updated_count = 0

        if inventory_type == "bar":
            for _, row in df.iterrows():
                name = str(row.get('name', row.get('item', 'Unknown'))).strip()
                existing = (await db.execute(
                    select(InventoryItem).where(InventoryItem.name == name, InventoryItem.is_active == True)
                )).scalar_one_or_none()
                if existing:
                    # Update fields if provided
                    if pd.notna(row.get('cost_per_unit', row.get('cost'))):
                        existing.cost_per_unit = float(row.get('cost_per_unit', row.get('cost', existing.cost_per_unit)))
                    if pd.notna(row.get('category', '')):
                        existing.category = str(row.get('category', existing.category or '')).strip() or existing.category
                    updated_count += 1
                else:
                    item = InventoryItem(
                        name=name,
                        category=str(row.get('category', '')).strip() or None,
                        subcategory=str(row.get('subcategory', '')).strip() or None,
                        location=str(row.get('location', '')).strip() or None,
                        section=str(row.get('section', '')).strip() or None,
                        bottle_size_ml=int(row.get('bottle_size_ml', row.get('size_ml', 0))) if pd.notna(row.get('bottle_size_ml', row.get('size_ml'))) else None,
                        cost_per_unit=float(row.get('cost_per_unit', row.get('cost', 0))) if pd.notna(row.get('cost_per_unit', row.get('cost'))) else 0.0
                    )
                    db.add(item)
                    imported_count += 1
        else:
            for _, row in df.iterrows():
                name = str(row.get('name', row.get('item', 'Unknown'))).strip()
                existing = (await db.execute(
                    select(KitchenInventoryItem).where(KitchenInventoryItem.name == name, KitchenInventoryItem.is_active == True)
                )).scalar_one_or_none()
                if existing:
                    if pd.notna(row.get('cost_per_unit', row.get('cost'))):
                        existing.cost_per_unit = float(row.get('cost_per_unit', row.get('cost', existing.cost_per_unit)))
                    if pd.notna(row.get('par_level', row.get('par'))):
                        existing.par_level = float(row.get('par_level', row.get('par', existing.par_level)))
                    updated_count += 1
                else:
                    item = KitchenInventoryItem(
                        name=name,
                        unit=str(row.get('unit', '')).strip() or None,
                        location=str(row.get('location', '')).strip() or None,
                        station=str(row.get('station', '')).strip() or None,
                        cost_per_unit=float(row.get('cost_per_unit', row.get('cost', 0))) if pd.notna(row.get('cost_per_unit', row.get('cost'))) else 0.0,
                        vendor=str(row.get('vendor', '')).strip() or None,
                        par_level=float(row.get('par_level', row.get('par', 0))) if pd.notna(row.get('par_level', row.get('par'))) else 0.0
                    )
                    db.add(item)
                    imported_count += 1

        await db.commit()
        return {"message": f"Imported {imported_count} new, updated {updated_count} existing items", "imported": imported_count, "updated": updated_count}
        
    except Exception as e:
        logger.error(f"Import error: {e}")
        raise HTTPException(status_code=400, detail=f"Error processing file: {str(e)}")


# ============================================================
# Receipt / Invoice OCR Parsing (Gemini Vision)
# ============================================================

import base64

class ReceiptParseResult(BaseModel):
    vendor: Optional[str] = None
    date: Optional[str] = None          # ISO date string YYYY-MM-DD
    items: List[dict] = []              # [{name, quantity, unit, unit_cost, total_cost, purchase_type}]
    subtotal: Optional[float] = None
    tax: Optional[float] = None
    total: Optional[float] = None
    notes: Optional[str] = None
    raw_text: Optional[str] = None      # for debugging

@api_router.post("/import/receipt")
async def parse_receipt(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_manager),
):
    """
    Accept a photo (JPEG/PNG/WEBP) or PDF of a receipt/invoice.
    Use Gemini Vision to extract line items, then save as Purchase records.
    Returns the parsed data for user review before confirming.
    """
    content = await file.read()
    filename = file.filename.lower() if file.filename else ""

    # Determine media type
    if filename.endswith((".jpg", ".jpeg")) or file.content_type == "image/jpeg":
        media_type = "image/jpeg"
    elif filename.endswith(".png") or file.content_type == "image/png":
        media_type = "image/png"
    elif filename.endswith(".webp") or file.content_type == "image/webp":
        media_type = "image/webp"
    elif filename.endswith(".pdf") or file.content_type == "application/pdf":
        media_type = "application/pdf"
    else:
        raise HTTPException(status_code=400, detail="Unsupported file type. Upload a photo (JPEG/PNG/WEBP) or PDF.")

    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large. Maximum 20MB.")

    # Build Gemini prompt
    prompt = """You are a receipt and invoice parser for a restaurant.

Analyze this receipt or invoice image and extract ALL line items.

Return ONLY valid JSON in this exact format, no markdown, no explanation:
{
  "vendor": "vendor/supplier name or null",
  "date": "YYYY-MM-DD or null",
  "items": [
    {
      "name": "product name",
      "quantity": 1.0,
      "unit": "bottle/case/lb/each/etc or null",
      "unit_cost": 0.00,
      "total_cost": 0.00,
      "purchase_type": "bar or kitchen or supply or other"
    }
  ],
  "subtotal": 0.00,
  "tax": 0.00,
  "total": 0.00,
  "notes": "any relevant notes or null"
}

Rules:
- purchase_type: "bar" for alcohol/beverages, "kitchen" for food ingredients, "supply" for cleaning/paper goods, "other" for everything else
- If a field is unknown, use null
- quantity defaults to 1 if not shown
- total_cost = quantity * unit_cost if not explicit
- Include every line item — do not skip any
- Date format must be YYYY-MM-DD
- All costs must be numbers (no currency symbols)"""

    try:
        _client = genai_client.Client(api_key=_GEMINI_API_KEY)

        # Pass image/PDF as base64 inline
        b64_data = base64.b64encode(content).decode("utf-8")
        image_part = genai_client.types.Part.from_bytes(data=base64.b64decode(b64_data), mime_type=media_type)

        response = _client.models.generate_content(
            model=os.environ.get("GEMINI_RECEIPT_MODEL", "gemini-2.0-flash-lite"),
            contents=[prompt, image_part],
        )
        raw = response.text.strip()

        # Strip markdown fences if present
        if raw.startswith("```"):
            raw = raw.split("\n", 1)[-1]
            if raw.endswith("```"):
                raw = raw.rsplit("```", 1)[0]
            raw = raw.strip()

        parsed = json.loads(raw)

        return {
            "success": True,
            "vendor": parsed.get("vendor"),
            "date": parsed.get("date"),
            "items": parsed.get("items", []),
            "subtotal": parsed.get("subtotal"),
            "tax": parsed.get("tax"),
            "total": parsed.get("total"),
            "notes": parsed.get("notes"),
        }

    except json.JSONDecodeError as e:
        logger.error(f"Receipt parse JSON error: {e} | raw: {raw[:200]}")
        raise HTTPException(status_code=422, detail="Could not parse receipt data. Try a clearer photo.")
    except Exception as e:
        logger.error(f"Receipt parse error: {e}")
        raise HTTPException(status_code=500, detail=f"Receipt parsing failed: {str(e)}")


@api_router.post("/import/receipt/confirm")
async def confirm_receipt_purchases(
    data: dict,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_manager),
):
    """
    Save confirmed receipt line items as Purchase records.
    Expects: {vendor, date, items: [{name, quantity, unit, total_cost, purchase_type}]}
    """
    items  = data.get("items", [])
    vendor = data.get("vendor", "Unknown Vendor")
    date_str = data.get("date")

    if not items:
        raise HTTPException(status_code=400, detail="No items to save")

    try:
        purchase_date = datetime.fromisoformat(date_str).replace(tzinfo=timezone.utc) if date_str else datetime.now(timezone.utc)
    except Exception:
        purchase_date = datetime.now(timezone.utc)

    saved = 0
    for item in items:
        name  = item.get("name", "Unknown Item").strip()
        if not name:
            continue
        p = Purchase(
            item_name=f"{name} ({vendor})" if vendor else name,
            item_type=item.get("purchase_type", "other"),
            purchase_type=item.get("purchase_type", "other"),
            quantity=float(item.get("quantity") or 1),
            total_cost=float(item.get("total_cost") or 0),
            date=purchase_date,
        )
        db.add(p)
        saved += 1

    await db.commit()
    return {"message": f"{saved} purchase{'s' if saved != 1 else ''} recorded from receipt", "saved": saved}

# Health check
@api_router.get("/")
async def root():
    return {"message": "Ops AI API", "status": "running"}

@api_router.get("/health")
async def health():
    return {"status": "healthy"}

# Admin endpoint to clear all data
class ClearDataConfirm(BaseModel):
    confirm: str  # must equal "DELETE_ALL"

@api_router.delete("/admin/clear-data")
async def clear_all_data(body: ClearDataConfirm, db: AsyncSession = Depends(get_db), admin: User = Depends(require_admin)):
    """Clear all data — requires explicit confirmation body: {"confirm": "DELETE_ALL"}"""
    if body.confirm != "DELETE_ALL":
        raise HTTPException(status_code=400, detail='Send {"confirm": "DELETE_ALL"} to confirm.')
    from sqlalchemy import delete
    logger.warning(f"AUDIT: /admin/clear-data triggered by user={admin.name!r} (id={admin.id}) at {datetime.now(timezone.utc).isoformat()}")
    
    # Clear counts first (foreign keys)
    await db.execute(delete(InventoryCount))
    await db.execute(delete(KitchenInventoryCount))
    
    # Clear items
    await db.execute(delete(InventoryItem))
    await db.execute(delete(KitchenInventoryItem))
    
    # Clear purchases and sales
    await db.execute(delete(Purchase))
    await db.execute(delete(Sale))
    
    # Clear menu data
    await db.execute(delete(MenuItemIngredient))
    await db.execute(delete(MenuItem))
    
    await db.commit()
    return {"message": "All data cleared successfully"}

# ── Scheduled Low-Stock Alert ────────────────────────────────────────────────

async def _scheduled_low_stock_check():
    """Run at 07:00 UTC daily. Send push alert if any items are critical/high."""
    if not VAPID_PRIVATE_KEY:
        return
    async with AsyncSessionLocal() as db:
        # Bar items — bulk latest counts
        bar_ts = (
            select(InventoryCount.item_id, func.max(InventoryCount.timestamp).label("max_ts"))
            .group_by(InventoryCount.item_id).subquery()
        )
        bar_ct = (
            select(InventoryCount.item_id, InventoryCount.level_percentage)
            .join(bar_ts, and_(InventoryCount.item_id == bar_ts.c.item_id,
                               InventoryCount.timestamp == bar_ts.c.max_ts)).subquery()
        )
        bar_rows = (await db.execute(
            select(InventoryItem.name, bar_ct.c.level_percentage)
            .outerjoin(bar_ct, InventoryItem.id == bar_ct.c.item_id)
            .where(InventoryItem.is_active == True)
        )).all()
        critical_bar = [name for name, pct in bar_rows if pct is not None and pct <= 10]

        # Kitchen items — bulk latest counts
        kit_ts = (
            select(KitchenInventoryCount.item_id, func.max(KitchenInventoryCount.timestamp).label("max_ts"))
            .group_by(KitchenInventoryCount.item_id).subquery()
        )
        kit_ct = (
            select(KitchenInventoryCount.item_id, KitchenInventoryCount.quantity)
            .join(kit_ts, and_(KitchenInventoryCount.item_id == kit_ts.c.item_id,
                               KitchenInventoryCount.timestamp == kit_ts.c.max_ts)).subquery()
        )
        kit_rows = (await db.execute(
            select(KitchenInventoryItem.name, KitchenInventoryItem.par_level, kit_ct.c.quantity)
            .outerjoin(kit_ct, KitchenInventoryItem.id == kit_ct.c.item_id)
            .where(KitchenInventoryItem.is_active == True)
        )).all()
        critical_kit = [name for name, par, qty in kit_rows
                        if qty is not None and par and par > 0 and (qty / par) <= 0.25]

        total = len(critical_bar) + len(critical_kit)
        if total == 0:
            logger.info("Low-stock check: no critical items.")
            return

        sample = (critical_bar + critical_kit)[:3]
        more = total - len(sample)
        body = f"{', '.join(sample)}" + (f" +{more} more" if more else "")
        title = f"⚠️ {total} item{'s' if total != 1 else ''} critically low"

        await _send_push_to_all(db, title=title, body=body, url="/reports?tab=low-stock")
        logger.info(f"Low-stock alert sent: {total} items — {body}")

# ── Push Notifications ────────────────────────────────────────────────────────

VAPID_PRIVATE_KEY = os.environ.get("VAPID_PRIVATE_KEY", "")
VAPID_PUBLIC_KEY = os.environ.get("VAPID_PUBLIC_KEY", "")
VAPID_CLAIMS = {"sub": os.environ.get("VAPID_MAILTO", "mailto:ops@opsai.app")}

class PushSubscriptionCreate(BaseModel):
    endpoint: str
    keys: dict  # {p256dh: str, auth: str}

@api_router.post("/push/subscribe")
async def push_subscribe(data: PushSubscriptionCreate, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    # Upsert by endpoint
    result = await db.execute(select(PushSubscription).where(PushSubscription.endpoint == data.endpoint))
    existing = result.scalar_one_or_none()
    if existing:
        existing.user_id = user.id
        existing.p256dh = data.keys["p256dh"]
        existing.auth = data.keys["auth"]
    else:
        sub = PushSubscription(
            user_id=user.id,
            endpoint=data.endpoint,
            p256dh=data.keys["p256dh"],
            auth=data.keys["auth"],
        )
        db.add(sub)
    await db.commit()
    return {"status": "subscribed"}

@api_router.delete("/push/subscribe")
async def push_unsubscribe(data: PushSubscriptionCreate, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    result = await db.execute(select(PushSubscription).where(PushSubscription.endpoint == data.endpoint))
    existing = result.scalar_one_or_none()
    if existing:
        await db.delete(existing)
        await db.commit()
    return {"status": "unsubscribed"}

@api_router.get("/push/vapid-key")
async def get_vapid_key():
    return {"publicKey": VAPID_PUBLIC_KEY}

class PushNotificationSend(BaseModel):
    title: str
    body: str
    url: Optional[str] = "/"
    user_ids: Optional[List[str]] = None  # None = broadcast to all

@api_router.post("/push/send")
async def send_push_notification(data: PushNotificationSend, db: AsyncSession = Depends(get_db), admin: User = Depends(require_admin)):
    query = select(PushSubscription)
    if data.user_ids:
        query = query.where(PushSubscription.user_id.in_(data.user_ids))
    result = await db.execute(query)
    subs = result.scalars().all()

    payload = json.dumps({"title": data.title, "body": data.body, "url": data.url})
    sent, failed = 0, 0
    stale_ids = []

    for sub in subs:
        try:
            webpush(
                subscription_info={"endpoint": sub.endpoint, "keys": {"p256dh": sub.p256dh, "auth": sub.auth}},
                data=payload,
                vapid_private_key=VAPID_PRIVATE_KEY,
                vapid_claims=VAPID_CLAIMS,
            )
            sent += 1
        except WebPushException as e:
            if e.response and e.response.status_code in (404, 410):
                stale_ids.append(sub.id)
            failed += 1
        except Exception:
            failed += 1

    # Clean up stale subscriptions
    if stale_ids:
        for sid in stale_ids:
            result = await db.execute(select(PushSubscription).where(PushSubscription.id == sid))
            stale = result.scalar_one_or_none()
            if stale:
                await db.delete(stale)
        await db.commit()

    return {"sent": sent, "failed": failed, "cleaned": len(stale_ids)}

@api_router.get("/ai/models")
async def list_ai_models(user: User = Depends(require_admin)):
    """List available Gemini models for debugging — uses httpx to call REST directly."""
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            r = await client.get(
                f"https://generativelanguage.googleapis.com/v1beta/models?key={_GEMINI_API_KEY}&pageSize=50"
            )
            if r.status_code != 200:
                return {"error": f"HTTP {r.status_code}: {r.text[:200]}"}
            data = r.json()
            models = data.get("models", [])
            generate = [m["name"] for m in models
                        if "generateContent" in m.get("supportedGenerationMethods", [])]
            return {"total": len(models), "generate_models": generate}
    except Exception as e:
        return {"error": str(e)[:300]}

@api_router.post("/push/test-low-stock")
async def trigger_low_stock_check(db: AsyncSession = Depends(get_db), admin: User = Depends(require_manager)):
    """Manually trigger the low-stock push alert (for testing). Manager+ only."""
    await _scheduled_low_stock_check()
    return {"message": "Low-stock check triggered"}

@api_router.get("/push/status")
async def push_status(db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    """Returns whether push is configured and how many subscriptions exist."""
    count_result = await db.execute(select(func.count(PushSubscription.id)))
    sub_count = count_result.scalar() or 0
    return {
        "configured": bool(VAPID_PRIVATE_KEY and VAPID_PUBLIC_KEY),
        "public_key": VAPID_PUBLIC_KEY if VAPID_PUBLIC_KEY else None,
        "subscriptions": sub_count,
    }

# Helper to send push from other parts of the app
async def _send_push_to_all(db: AsyncSession, title: str, body: str, url: str = "/"):
    if not VAPID_PRIVATE_KEY:
        return
    result = await db.execute(select(PushSubscription))
    subs = result.scalars().all()
    payload = json.dumps({"title": title, "body": body, "url": url})
    for sub in subs:
        try:
            webpush(
                subscription_info={"endpoint": sub.endpoint, "keys": {"p256dh": sub.p256dh, "auth": sub.auth}},
                data=payload,
                vapid_private_key=VAPID_PRIVATE_KEY,
                vapid_claims=VAPID_CLAIMS,
            )
        except Exception:
            pass

# (router and CORS registered at end of file)



# ============================================================
# Reports & Analytics Endpoints
# ============================================================

class WasteLogCreate(BaseModel):
    item_name: str
    item_type: str                  # bar | kitchen
    reason: str                     # waste | comp | spill | breakage | expired | other
    quantity: float = 1.0
    unit: Optional[str] = None
    estimated_cost: float = 0.0
    notes: Optional[str] = None
    date: Optional[datetime] = None  # defaults to now if not supplied

# ── Waste Log ──────────────────────────────────────────────

@api_router.post("/reports/waste")
async def log_waste(data: WasteLogCreate, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    log_date = data.date if data.date else datetime.now(timezone.utc)
    if log_date.tzinfo is None:
        log_date = log_date.replace(tzinfo=timezone.utc)
    entry = WasteLog(
        item_name=data.item_name, item_type=data.item_type, reason=data.reason,
        quantity=data.quantity, unit=data.unit, estimated_cost=data.estimated_cost,
        notes=data.notes, logged_by=user.id, date=log_date,
    )
    db.add(entry)
    await db.commit()
    return {"message": "Waste logged", "id": entry.id}

@api_router.get("/reports/waste")
async def get_waste_logs(days: int = 30, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    start = datetime.now(timezone.utc) - timedelta(days=days)
    result = await db.execute(
        select(WasteLog).where(WasteLog.date >= start).order_by(desc(WasteLog.date))
    )
    logs = result.scalars().all()
    return [
        {
            "id": l.id, "item_name": l.item_name, "item_type": l.item_type,
            "reason": l.reason, "quantity": l.quantity, "unit": l.unit,
            "estimated_cost": l.estimated_cost, "notes": l.notes,
            "date": l.date.isoformat(),
        }
        for l in logs
    ]

@api_router.delete("/reports/waste/{log_id}")
async def delete_waste_log(log_id: str, db: AsyncSession = Depends(get_db), user: User = Depends(require_manager)):
    result = await db.execute(select(WasteLog).where(WasteLog.id == log_id))
    entry = result.scalar_one_or_none()
    if not entry:
        raise HTTPException(status_code=404, detail="Log not found")
    await db.delete(entry)
    await db.commit()
    return {"message": "Deleted"}

# ── Sales Report ───────────────────────────────────────────

@api_router.get("/reports/sales")
async def report_sales(days: int = 30, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    start = datetime.now(timezone.utc) - timedelta(days=days)
    result = await db.execute(
        select(Sale).where(Sale.date >= start).order_by(Sale.date)
    )
    sales = result.scalars().all()
    rows = [{"date": s.date.strftime("%Y-%m-%d"), "total": s.total_sales, "bar": s.bar_sales, "food": s.food_sales} for s in sales]
    total  = sum(r["total"] for r in rows)
    bar    = sum(r["bar"]   for r in rows)
    food   = sum(r["food"]  for r in rows)
    avg    = round(total / len(rows), 2) if rows else 0
    return {
        "period_days": days, "rows": rows,
        "summary": {"total_sales": round(total,2), "bar_sales": round(bar,2), "food_sales": round(food,2), "avg_daily_sales": avg, "days_with_data": len(rows)},
    }

# ── Liquor / Pour Cost Report ──────────────────────────────

@api_router.get("/reports/pour-cost")
async def report_pour_cost(days: int = 30, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    start = datetime.now(timezone.utc) - timedelta(days=days)

    # Bar purchases
    purch_result = await db.execute(
        select(Purchase).where(and_(Purchase.date >= start, Purchase.purchase_type == "bar")).order_by(Purchase.date)
    )
    purchases = purch_result.scalars().all()

    # Bar sales
    sales_result = await db.execute(
        select(func.sum(Sale.bar_sales)).where(Sale.date >= start)
    )
    bar_sales = sales_result.scalar() or 0

    # Waste cost (bar)
    waste_result = await db.execute(
        select(func.sum(WasteLog.estimated_cost)).where(and_(WasteLog.date >= start, WasteLog.item_type == "bar"))
    )
    waste_cost = waste_result.scalar() or 0

    total_cogs   = sum(p.total_cost for p in purchases)
    pour_cost_pct = round((total_cogs / bar_sales * 100) if bar_sales > 0 else 0, 1)
    target        = float(os.environ.get("TARGET_POUR_COST_PCT", "20.0"))
    variance      = round(pour_cost_pct - target, 1)

    # By category
    cat_result = await db.execute(
        select(Purchase.item_name, func.sum(Purchase.total_cost))
        .where(and_(Purchase.date >= start, Purchase.purchase_type == "bar"))
        .group_by(Purchase.item_name)
        .order_by(desc(func.sum(Purchase.total_cost)))
    )
    by_item = [{"name": r[0], "cost": round(r[1], 2)} for r in cat_result.all()]

    return {
        "period_days": days,
        "bar_sales": round(bar_sales, 2),
        "bar_cogs": round(total_cogs, 2),
        "waste_cost": round(waste_cost, 2),
        "pour_cost_pct": pour_cost_pct,
        "target_pct": target,
        "variance": variance,
        "status": "over" if variance > 0 else "on_target" if variance == 0 else "under",
        "top_purchases": by_item[:10],
    }

# ── Food Cost Report ───────────────────────────────────────

@api_router.get("/reports/food-cost")
async def report_food_cost(days: int = 30, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    start = datetime.now(timezone.utc) - timedelta(days=days)

    purch_result = await db.execute(
        select(Purchase).where(and_(Purchase.date >= start, Purchase.purchase_type == "kitchen"))
    )
    purchases = purch_result.scalars().all()

    sales_result = await db.execute(select(func.sum(Sale.food_sales)).where(Sale.date >= start))
    food_sales = sales_result.scalar() or 0

    waste_result = await db.execute(
        select(func.sum(WasteLog.estimated_cost)).where(and_(WasteLog.date >= start, WasteLog.item_type == "kitchen"))
    )
    waste_cost = waste_result.scalar() or 0

    total_cogs    = sum(p.total_cost for p in purchases)
    food_cost_pct = round((total_cogs / food_sales * 100) if food_sales > 0 else 0, 1)
    target        = float(os.environ.get("TARGET_FOOD_COST_PCT", "30.0"))
    variance      = round(food_cost_pct - target, 1)

    menu_result = await db.execute(
        select(MenuItem).where(MenuItem.is_active == True).options(selectinload(MenuItem.ingredients))
    )
    menu_items = menu_result.scalars().all()
    menu_rows = []
    for item in menu_items:
        cost = sum(i.quantity_used * i.cost_per_unit for i in item.ingredients)
        pct  = round((cost / item.price * 100) if item.price > 0 else 0, 1)
        menu_rows.append({"name": item.name, "category": item.category, "price": item.price, "cost": round(cost,2), "cost_pct": pct, "margin": round(item.price - cost, 2)})
    menu_rows.sort(key=lambda x: x["cost_pct"], reverse=True)

    return {
        "period_days": days,
        "food_sales": round(food_sales, 2),
        "food_cogs": round(total_cogs, 2),
        "waste_cost": round(waste_cost, 2),
        "food_cost_pct": food_cost_pct,
        "target_pct": target,
        "variance": variance,
        "status": "over" if variance > 0 else "on_target" if variance == 0 else "under",
        "menu_items": menu_rows,
    }

# ── Inventory Variance Report ──────────────────────────────

@api_router.get("/reports/inventory-variance")
async def report_inventory_variance(db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    """Compare the two most recent counts per bar item to show movement/shrinkage."""
    items_result = await db.execute(select(InventoryItem).where(InventoryItem.is_active == True))
    items = items_result.scalars().all()

    rows = []
    for item in items:
        counts_result = await db.execute(
            select(InventoryCount.level_percentage, InventoryCount.timestamp)
            .where(InventoryCount.item_id == item.id)
            .order_by(desc(InventoryCount.timestamp))
            .limit(2)
        )
        counts = counts_result.all()
        if len(counts) < 2:
            continue
        latest, previous = counts[0], counts[1]
        change  = latest[0] - previous[0]
        # Estimate bottle value lost
        cost_impact = round(abs(change) / 100 * (item.cost_per_unit or 0), 2)
        rows.append({
            "name": item.name, "location": item.location, "category": item.category,
            "previous_pct": previous[0], "current_pct": latest[0],
            "change_pct": change,
            "previous_date": previous[1].strftime("%Y-%m-%d %H:%M"),
            "current_date":  latest[1].strftime("%Y-%m-%d %H:%M"),
            "cost_impact": cost_impact,
            "flag": "shrinkage" if change < -25 else ("low" if latest[0] <= 25 else "ok"),
        })

    rows.sort(key=lambda x: x["change_pct"])
    total_shrinkage_cost = round(sum(r["cost_impact"] for r in rows if r["change_pct"] < -25), 2)
    low_stock = [r for r in rows if r["flag"] in ("shrinkage", "low")]

    return {
        "items": rows,
        "summary": {
            "total_items": len(rows),
            "low_stock_count": len(low_stock),
            "shrinkage_items": len([r for r in rows if r["flag"] == "shrinkage"]),
            "estimated_shrinkage_cost": total_shrinkage_cost,
        }
    }

# ── Low Stock Report ───────────────────────────────────────

@api_router.get("/reports/low-stock")
async def report_low_stock(db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    """All bar + kitchen items at or below par / 25% threshold."""
    # Bar: bulk fetch latest count per item via MAX(timestamp) subquery — no N+1
    bar_ts_sub = (
        select(InventoryCount.item_id, func.max(InventoryCount.timestamp).label("max_ts"))
        .group_by(InventoryCount.item_id).subquery()
    )
    bar_count_sub = (
        select(InventoryCount.item_id, InventoryCount.level_percentage, InventoryCount.timestamp)
        .join(bar_ts_sub, and_(
            InventoryCount.item_id == bar_ts_sub.c.item_id,
            InventoryCount.timestamp == bar_ts_sub.c.max_ts,
        )).subquery()
    )
    bar_rows = (await db.execute(
        select(InventoryItem, bar_count_sub.c.level_percentage, bar_count_sub.c.timestamp)
        .outerjoin(bar_count_sub, InventoryItem.id == bar_count_sub.c.item_id)
        .where(InventoryItem.is_active == True)
    )).all()

    low_bar = []
    for item, level_pct, counted_at in bar_rows:
        if level_pct is not None and level_pct <= 25:
            low_bar.append({
                "name": item.name, "type": "bar", "location": item.location,
                "category": item.category, "level_pct": level_pct,
                "last_counted": counted_at.strftime("%Y-%m-%d %H:%M") if counted_at else None,
                "cost_per_unit": item.cost_per_unit,
                "urgency": "critical" if level_pct == 0 else ("high" if level_pct <= 10 else "medium"),
            })

    # Kitchen: same pattern
    kit_ts_sub = (
        select(KitchenInventoryCount.item_id, func.max(KitchenInventoryCount.timestamp).label("max_ts"))
        .group_by(KitchenInventoryCount.item_id).subquery()
    )
    kit_count_sub = (
        select(KitchenInventoryCount.item_id, KitchenInventoryCount.quantity, KitchenInventoryCount.timestamp)
        .join(kit_ts_sub, and_(
            KitchenInventoryCount.item_id == kit_ts_sub.c.item_id,
            KitchenInventoryCount.timestamp == kit_ts_sub.c.max_ts,
        )).subquery()
    )
    kit_rows = (await db.execute(
        select(KitchenInventoryItem, kit_count_sub.c.quantity, kit_count_sub.c.timestamp)
        .outerjoin(kit_count_sub, KitchenInventoryItem.id == kit_count_sub.c.item_id)
        .where(KitchenInventoryItem.is_active == True)
    )).all()

    low_kitchen = []
    for item, quantity, counted_at in kit_rows:
        if quantity is not None and item.par_level > 0 and quantity < item.par_level:
            pct_of_par = round(quantity / item.par_level * 100)
            low_kitchen.append({
                "name": item.name, "type": "kitchen", "location": item.location,
                "station": item.station, "quantity": quantity, "par_level": item.par_level,
                "unit": item.unit, "pct_of_par": pct_of_par,
                "last_counted": counted_at.strftime("%Y-%m-%d %H:%M") if counted_at else None,
                "cost_per_unit": item.cost_per_unit,
                "urgency": "critical" if quantity == 0 else ("high" if pct_of_par <= 25 else "medium"),
            })

    all_low = sorted(low_bar + low_kitchen, key=lambda x: {"critical": 0, "high": 1, "medium": 2}[x["urgency"]])
    return {
        "items": all_low,
        "summary": {
            "total_low": len(all_low),
            "critical": len([i for i in all_low if i["urgency"] == "critical"]),
            "high": len([i for i in all_low if i["urgency"] == "high"]),
            "medium": len([i for i in all_low if i["urgency"] == "medium"]),
        }
    }

# ── Waste Summary Report ───────────────────────────────────

@api_router.get("/reports/waste-summary")
async def report_waste_summary(days: int = 30, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    start = datetime.now(timezone.utc) - timedelta(days=days)
    result = await db.execute(
        select(WasteLog.reason, WasteLog.item_type, func.count(WasteLog.id), func.sum(WasteLog.estimated_cost))
        .where(WasteLog.date >= start)
        .group_by(WasteLog.reason, WasteLog.item_type)
    )
    rows = result.all()

    by_reason = {}
    for reason, itype, count, cost in rows:
        if reason not in by_reason:
            by_reason[reason] = {"reason": reason, "count": 0, "cost": 0.0, "bar": 0.0, "kitchen": 0.0}
        by_reason[reason]["count"] += count
        by_reason[reason]["cost"]  += cost or 0
        by_reason[reason][itype]   = round((by_reason[reason].get(itype, 0) or 0) + (cost or 0), 2)

    detail_result = await db.execute(
        select(WasteLog).where(WasteLog.date >= start).order_by(desc(WasteLog.date)).limit(50)
    )
    recent = detail_result.scalars().all()

    total_cost = sum(r.estimated_cost or 0 for r in recent)
    return {
        "period_days": days,
        "total_waste_cost": round(total_cost, 2),
        "by_reason": sorted(by_reason.values(), key=lambda x: x["cost"], reverse=True),
        "recent_entries": [
            {"date": r.date.strftime("%Y-%m-%d"), "item": r.item_name, "reason": r.reason,
             "quantity": r.quantity, "unit": r.unit, "cost": r.estimated_cost, "notes": r.notes}
            for r in recent
        ],
    }

# ============================================================
# Toast POS Integration
# ============================================================

# Toast uses the same base URL for both auth and API in production
# Override individually for sandbox testing
TOAST_AUTH_BASE = os.environ.get("TOAST_AUTH_URL", os.environ.get("TOAST_API_BASE_URL", "https://ws-api.toasttab.com"))
TOAST_API_BASE  = os.environ.get("TOAST_API_URL",  os.environ.get("TOAST_API_BASE_URL", "https://ws-api.toasttab.com"))
# For sandbox: set TOAST_API_BASE_URL=https://ws-sandbox-api.eng.toasttab.com
# Or set TOAST_AUTH_URL and TOAST_API_URL individually

# Pydantic schemas
class ToastConnectRequest(BaseModel):
    client_id: str
    client_secret: str
    restaurant_guid: str
    restaurant_name: Optional[str] = None

class ToastStatusResponse(BaseModel):
    is_connected: bool
    restaurant_name: Optional[str]
    restaurant_guid: Optional[str]
    last_synced_at: Optional[str]
    last_sync_status: Optional[str]
    last_sync_message: Optional[str]
    connected_at: Optional[str]

async def _get_toast_record(db: AsyncSession) -> Optional[ToastIntegration]:
    """Return the single ToastIntegration row (singleton per deployment)."""
    result = await db.execute(select(ToastIntegration).limit(1))
    return result.scalar_one_or_none()

async def _refresh_toast_token(record: ToastIntegration, db: AsyncSession) -> bool:
    """Exchange client credentials for a new access token using Toast OAuth."""
    try:
        async with httpx.AsyncClient(timeout=15) as client:
            resp = await client.post(
                f"{TOAST_AUTH_BASE}/authentication/v1/authentication/login",
                json={
                    "clientId": record.client_id,
                    "clientSecret": record.client_secret,
                    "userAccessType": "TOAST_MACHINE_CLIENT",
                },
                headers={"Content-Type": "application/json"},
            )
        if resp.status_code != 200:
            logger.error(f"Toast token refresh failed: {resp.status_code} {resp.text}")
            return False
        data = resp.json()
        token = data.get("token", {})
        record.access_token  = token.get("accessToken")
        record.refresh_token = token.get("refreshToken")
        expires_in = token.get("expiresIn", 3600)
        record.token_expires_at = datetime.now(timezone.utc) + timedelta(seconds=expires_in)
        record.is_connected = True
        await db.commit()
        return True
    except Exception as e:
        logger.error(f"Toast token refresh exception: {e}")
        return False

async def _ensure_valid_token(record: ToastIntegration, db: AsyncSession) -> bool:
    """Ensure the access token is valid, refreshing if needed."""
    if not record.client_id or not record.client_secret:
        return False
    if not record.access_token:
        return await _refresh_toast_token(record, db)
    if record.token_expires_at:
        # Refresh 5 min before expiry
        if datetime.now(timezone.utc) >= record.token_expires_at - timedelta(minutes=5):
            return await _refresh_toast_token(record, db)
    return True

# ── Endpoints ──────────────────────────────────────────────

@api_router.post("/integrations/toast/connect")
async def toast_connect(
    data: ToastConnectRequest,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
):
    """Save Toast credentials and verify they work by fetching a token."""
    record = await _get_toast_record(db)
    if not record:
        record = ToastIntegration()
        db.add(record)

    record.client_id       = data.client_id
    record.client_secret   = data.client_secret
    record.restaurant_guid = data.restaurant_guid
    record.restaurant_name = data.restaurant_name or data.restaurant_guid
    record.is_connected    = False
    await db.flush()

    ok = await _refresh_toast_token(record, db)
    if not ok:
        raise HTTPException(status_code=400, detail="Could not authenticate with Toast. Check your Client ID, Client Secret, and Restaurant GUID.")

    record.connected_at = datetime.now(timezone.utc)
    await db.commit()
    logger.info(f"Toast connected for restaurant: {record.restaurant_name}")
    return {"message": "Toast connected successfully", "restaurant": record.restaurant_name}


@api_router.get("/integrations/toast/status", response_model=ToastStatusResponse)
async def toast_status(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Return current connection status."""
    record = await _get_toast_record(db)
    if not record:
        return ToastStatusResponse(
            is_connected=False, restaurant_name=None, restaurant_guid=None,
            last_synced_at=None, last_sync_status=None, last_sync_message=None, connected_at=None,
        )
    return ToastStatusResponse(
        is_connected=record.is_connected,
        restaurant_name=record.restaurant_name,
        restaurant_guid=record.restaurant_guid,
        last_synced_at=record.last_synced_at.isoformat() if record.last_synced_at else None,
        last_sync_status=record.last_sync_status,
        last_sync_message=record.last_sync_message,
        connected_at=record.connected_at.isoformat() if record.connected_at else None,
    )


@api_router.post("/integrations/toast/disconnect")
async def toast_disconnect(
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
):
    """Clear stored tokens and mark as disconnected."""
    record = await _get_toast_record(db)
    if record:
        record.access_token    = None
        record.refresh_token   = None
        record.is_connected    = False
        record.last_sync_status = None
        await db.commit()
    return {"message": "Toast disconnected"}


@api_router.post("/integrations/toast/sync")
async def toast_sync(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_manager),
):
    """Pull yesterday's sales from Toast and upsert into the sales table."""
    record = await _get_toast_record(db)
    if not record or not record.is_connected:
        raise HTTPException(status_code=400, detail="Toast is not connected")

    ok = await _ensure_valid_token(record, db)
    if not ok:
        record.last_sync_status  = "error"
        record.last_sync_message = "Token refresh failed — check credentials"
        await db.commit()
        raise HTTPException(status_code=401, detail="Toast authentication failed. Reconnect in Integrations.")

    # Pull orders for yesterday (local midnight → midnight)
    today     = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    yesterday = today - timedelta(days=1)

    try:
        async with httpx.AsyncClient(timeout=30) as client:
            resp = await client.get(
                f"{TOAST_API_BASE}/orders/v2/ordersBulk",
                headers={
                    "Authorization": f"Bearer {record.access_token}",
                    "Toast-Restaurant-External-ID": record.restaurant_guid,
                },
                params={
                    "startDate": yesterday.strftime("%Y-%m-%dT%H:%M:%S.000+0000"),
                    "endDate":   today.strftime("%Y-%m-%dT%H:%M:%S.000+0000"),
                    "pageSize":  500,
                },
            )

        if resp.status_code == 401:
            record.last_sync_status  = "error"
            record.last_sync_message = "Unauthorized — token may have expired"
            await db.commit()
            raise HTTPException(status_code=401, detail="Toast returned 401 — try reconnecting")

        if resp.status_code != 200:
            msg = f"Toast API error {resp.status_code}"
            record.last_sync_status  = "error"
            record.last_sync_message = msg
            await db.commit()
            raise HTTPException(status_code=502, detail=msg)

        orders = resp.json()  # list of order objects

        # Aggregate totals
        total_sales = 0.0
        bar_sales   = 0.0
        food_sales  = 0.0

        # Toast revenue centers: configurable via TOAST_BAR_REVENUE_CENTER_GUIDS env var
        # Fallback: name-based heuristic (set env var for reliable classification)
        _bar_rc_guids_raw = os.environ.get("TOAST_BAR_REVENUE_CENTER_GUIDS", "")
        _bar_rc_guids = {g.strip().lower() for g in _bar_rc_guids_raw.split(",") if g.strip()}

        for order in orders:
            checks = order.get("checks", [])
            for check in checks:
                amount = check.get("totalAmount", 0.0) or 0.0
                total_sales += amount
                rc = order.get("revenueCenter") or {}
                rc_guid = (rc.get("guid") or "").lower()
                rc_name = (rc.get("name") or "").lower()
                # GUID match takes priority; fall back to name heuristic
                if (_bar_rc_guids and rc_guid in _bar_rc_guids) or \
                   (not _bar_rc_guids and ("bar" in rc_name or "beverage" in rc_name or "drink" in rc_name)):
                    bar_sales += amount
                else:
                    food_sales += amount

        # Upsert into sales table for yesterday
        existing = await db.execute(
            select(Sale).where(
                Sale.date >= yesterday,
                Sale.date < today,
            )
        )
        existing_sale = existing.scalar_one_or_none()

        if existing_sale:
            existing_sale.total_sales = round(total_sales, 2)
            existing_sale.bar_sales   = round(bar_sales, 2)
            existing_sale.food_sales  = round(food_sales, 2)
        else:
            new_sale = Sale(
                date=yesterday,
                total_sales=round(total_sales, 2),
                bar_sales=round(bar_sales, 2),
                food_sales=round(food_sales, 2),
            )
            db.add(new_sale)

        record.last_synced_at    = datetime.now(timezone.utc)
        record.last_sync_status  = "success"
        record.last_sync_message = f"Synced {len(orders)} orders | ${total_sales:,.2f} total sales"
        await db.commit()

        return {
            "message": "Sync complete",
            "orders_processed": len(orders),
            "total_sales": round(total_sales, 2),
            "bar_sales":   round(bar_sales, 2),
            "food_sales":  round(food_sales, 2),
            "date": yesterday.strftime("%Y-%m-%d"),
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Toast sync error: {e}")
        record.last_sync_status  = "error"
        record.last_sync_message = str(e)
        await db.commit()
        raise HTTPException(status_code=500, detail=f"Sync failed: {str(e)}")



# ============================================================
# Report Export — Excel & PDF
# ============================================================

import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Excel helpers ─────────────────────────────────────────
_GOLD   = "FFD4A017"
_DARK   = "FF0A0A12"
_HEADER = "FF1A1A2E"
_TEXT   = "FFF5F5F0"
_MUTED  = "FF8E8E9F"

def _xl_header_style(cell, *, gold=False):
    cell.font      = Font(bold=True, color=_GOLD if gold else _TEXT, size=11)
    cell.fill      = PatternFill("solid", fgColor=_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = Border(bottom=Side(style="thin", color=_GOLD if gold else "FF2B2B4A"))

def _xl_title_row(ws, title: str, ncols: int):
    ws.append([title] + [""] * (ncols - 1))
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=ncols)
    cell = ws.cell(row=ws.max_row, column=1)
    cell.font      = Font(bold=True, color=_GOLD, size=13)
    cell.fill      = PatternFill("solid", fgColor=_DARK)
    cell.alignment = Alignment(horizontal="left", vertical="center")

def _xl_auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

def _wb_to_stream(wb) -> io.BytesIO:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ── PDF helpers ───────────────────────────────────────────
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch

_PDF_GOLD   = colors.HexColor("#D4A017")
_PDF_DARK   = colors.HexColor("#0A0A12")
_PDF_PANEL  = colors.HexColor("#1A1A2E")
_PDF_MUTED  = colors.HexColor("#8E8E9F")
_PDF_WHITE  = colors.white

def _pdf_doc(title: str, rows: list, col_headers: list, col_widths: list = None, summary_rows: list = None) -> io.BytesIO:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, leftMargin=0.5*inch, rightMargin=0.5*inch,
                            topMargin=0.5*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("OpsTitle", fontSize=16, textColor=_PDF_GOLD, fontName="Helvetica-Bold", spaceAfter=4)
    sub_style   = ParagraphStyle("OpsSub",   fontSize=9,  textColor=_PDF_MUTED, fontName="Helvetica", spaceAfter=12)

    elems = [
        Paragraph("OPS AI", title_style),
        Paragraph(title, sub_style),
    ]

    if summary_rows:
        sum_data = summary_rows
        sum_table = Table(sum_data, colWidths=[3*inch, 2*inch])
        sum_table.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,-1), _PDF_PANEL),
            ("TEXTCOLOR",  (0,0), (0,-1), _PDF_MUTED),
            ("TEXTCOLOR",  (1,0), (1,-1), _PDF_WHITE),
            ("FONTNAME",   (0,0), (-1,-1), "Helvetica"),
            ("FONTSIZE",   (0,0), (-1,-1), 9),
            ("ROWBACKGROUNDS", (0,0), (-1,-1), [_PDF_PANEL, colors.HexColor("#12121e")]),
            ("GRID",       (0,0), (-1,-1), 0.5, colors.HexColor("#2B2B4A")),
            ("TOPPADDING", (0,0), (-1,-1), 5),
            ("BOTTOMPADDING", (0,0), (-1,-1), 5),
        ]))
        elems += [sum_table, Spacer(1, 12)]

    if rows:
        page_width = letter[0] - inch
        if col_widths is None:
            col_widths = [page_width / len(col_headers)] * len(col_headers)

        data = [col_headers] + rows
        tbl  = Table(data, colWidths=col_widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            # Header row
            ("BACKGROUND",  (0,0), (-1,0), _PDF_DARK),
            ("TEXTCOLOR",   (0,0), (-1,0), _PDF_GOLD),
            ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",    (0,0), (-1,0), 9),
            ("ALIGN",       (0,0), (-1,0), "CENTER"),
            # Body
            ("FONTNAME",    (0,1), (-1,-1), "Helvetica"),
            ("FONTSIZE",    (0,1), (-1,-1), 8),
            ("TEXTCOLOR",   (0,1), (-1,-1), _PDF_WHITE),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [_PDF_PANEL, colors.HexColor("#12121e")]),
            ("GRID",        (0,0), (-1,-1), 0.5, colors.HexColor("#2B2B4A")),
            ("TOPPADDING",  (0,0), (-1,-1), 4),
            ("BOTTOMPADDING",(0,0), (-1,-1), 4),
        ]))
        elems.append(tbl)

    doc.build(elems)
    buf.seek(0)
    return buf

# ── Export endpoint ───────────────────────────────────────

@api_router.get("/reports/export")
async def export_report(
    report: str,           # sales | pour-cost | food-cost | waste | low-stock | variance
    format: str = "xlsx",  # xlsx | pdf
    days: int = 30,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    start = datetime.now(timezone.utc) - timedelta(days=days)
    ts    = datetime.now(timezone.utc).strftime("%Y%m%d")
    fmt   = format.lower()
    if fmt not in ("xlsx", "pdf"):
        raise HTTPException(status_code=400, detail="format must be xlsx or pdf")

    # ── SALES ────────────────────────────────────────────────
    if report == "sales":
        result = await db.execute(select(Sale).where(Sale.date >= start).order_by(Sale.date))
        sales  = result.scalars().all()
        rows   = [{"date": s.date.strftime("%Y-%m-%d"), "total": s.total_sales, "bar": s.bar_sales, "food": s.food_sales} for s in sales]
        total  = sum(r["total"] for r in rows)
        bar    = sum(r["bar"]   for r in rows)
        food   = sum(r["food"]  for r in rows)

        if fmt == "xlsx":
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sales"
            ws.sheet_view.showGridLines = False
            _xl_title_row(ws, f"Sales Report — Last {days} Days", 4)
            ws.append([])
            headers = ["Date", "Total Sales", "Bar Sales", "Food Sales"]
            ws.append(headers)
            for i, h in enumerate(headers, 1):
                _xl_header_style(ws.cell(ws.max_row, i), gold=(i == 1))
            for r in rows:
                ws.append([r["date"], r["total"], r["bar"], r["food"]])
            ws.append([])
            ws.append(["TOTAL", round(total,2), round(bar,2), round(food,2)])
            for i in range(1, 5):
                c = ws.cell(ws.max_row, i)
                c.font = Font(bold=True, color=_GOLD)
            _xl_auto_width(ws)
            filename = f"sales_{ts}.xlsx"
            return StreamingResponse(_wb_to_stream(wb), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                     headers={"Content-Disposition": f"attachment; filename={filename}"})
        else:
            summary = [["Period", f"Last {days} days"], ["Total Sales", f"${total:,.2f}"], ["Bar Sales", f"${bar:,.2f}"], ["Food Sales", f"${food:,.2f}"]]
            data_rows = [[r["date"], f"${r['total']:,.2f}", f"${r['bar']:,.2f}", f"${r['food']:,.2f}"] for r in rows]
            buf = _pdf_doc(f"Sales Report — Last {days} Days", data_rows, ["Date","Total","Bar","Food"], summary_rows=summary)
            return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=sales_{ts}.pdf"})

    # ── POUR COST ────────────────────────────────────────────
    elif report == "pour-cost":
        purch_result = await db.execute(select(Purchase).where(and_(Purchase.date >= start, Purchase.purchase_type == "bar")).order_by(Purchase.date))
        purchases = purch_result.scalars().all()
        bar_sales = (await db.execute(select(func.sum(Sale.bar_sales)).where(Sale.date >= start))).scalar() or 0
        total_cogs = sum(p.total_cost for p in purchases)
        pour_pct   = round((total_cogs / bar_sales * 100) if bar_sales > 0 else 0, 1)
        target     = float(os.environ.get("TARGET_POUR_COST_PCT", "20.0"))

        cat_result = await db.execute(
            select(Purchase.item_name, func.sum(Purchase.total_cost)).where(and_(Purchase.date >= start, Purchase.purchase_type == "bar"))
            .group_by(Purchase.item_name).order_by(desc(func.sum(Purchase.total_cost)))
        )
        by_item = cat_result.all()

        if fmt == "xlsx":
            wb = openpyxl.Workbook()
            ws = wb.active; ws.title = "Pour Cost"; ws.sheet_view.showGridLines = False
            _xl_title_row(ws, f"Pour Cost Report — Last {days} Days", 3)
            ws.append([])
            ws.append(["Bar Sales", f"${bar_sales:,.2f}", ""]); ws.append(["COGS (Bar Purchases)", f"${total_cogs:,.2f}", ""])
            ws.append(["Pour Cost %", f"{pour_pct}%", ""]); ws.append(["Target", f"{target}%", ""])
            ws.append(["Variance", f"{round(pour_pct-target,1)}%", "OVER" if pour_pct > target else "ON TARGET"])
            ws.append([]); ws.append(["Item", "Total Spend", "% of COGS"])
            for i in range(1,4): _xl_header_style(ws.cell(ws.max_row, i), gold=(i==1))
            for name, cost in by_item:
                pct_of = round(cost / total_cogs * 100, 1) if total_cogs else 0
                ws.append([name, round(cost,2), f"{pct_of}%"])
            _xl_auto_width(ws)
            return StreamingResponse(_wb_to_stream(wb), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                     headers={"Content-Disposition": f"attachment; filename=pour_cost_{ts}.xlsx"})
        else:
            summary = [["Period", f"Last {days} days"], ["Bar Sales", f"${bar_sales:,.2f}"], ["COGS", f"${total_cogs:,.2f}"],
                       ["Pour Cost %", f"{pour_pct}%"], ["Target", f"{target}%"], ["Variance", f"{round(pour_pct-target,1)}%"]]
            data_rows = [[n, f"${c:,.2f}", f"{round(c/total_cogs*100,1) if total_cogs else 0}%"] for n, c in by_item]
            buf = _pdf_doc(f"Pour Cost Report — Last {days} Days", data_rows, ["Item","Spend","% COGS"], summary_rows=summary)
            return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=pour_cost_{ts}.pdf"})

    # ── FOOD COST ────────────────────────────────────────────
    elif report == "food-cost":
        purch_result = await db.execute(select(Purchase).where(and_(Purchase.date >= start, Purchase.purchase_type == "kitchen")).order_by(Purchase.date))
        purchases  = purch_result.scalars().all()
        food_sales = (await db.execute(select(func.sum(Sale.food_sales)).where(Sale.date >= start))).scalar() or 0
        total_cogs = sum(p.total_cost for p in purchases)
        food_pct   = round((total_cogs / food_sales * 100) if food_sales > 0 else 0, 1)
        target     = float(os.environ.get("TARGET_FOOD_COST_PCT", "30.0"))

        cat_result = await db.execute(
            select(Purchase.item_name, func.sum(Purchase.total_cost)).where(and_(Purchase.date >= start, Purchase.purchase_type == "kitchen"))
            .group_by(Purchase.item_name).order_by(desc(func.sum(Purchase.total_cost)))
        )
        by_item = cat_result.all()

        if fmt == "xlsx":
            wb = openpyxl.Workbook()
            ws = wb.active; ws.title = "Food Cost"; ws.sheet_view.showGridLines = False
            _xl_title_row(ws, f"Food Cost Report — Last {days} Days", 3)
            ws.append([]); ws.append(["Food Sales", f"${food_sales:,.2f}", ""])
            ws.append(["COGS (Kitchen Purchases)", f"${total_cogs:,.2f}", ""])
            ws.append(["Food Cost %", f"{food_pct}%", ""]); ws.append(["Target", f"{target}%", ""])
            ws.append(["Variance", f"{round(food_pct-target,1)}%", "OVER" if food_pct > target else "ON TARGET"])
            ws.append([]); ws.append(["Item", "Total Spend", "% of COGS"])
            for i in range(1,4): _xl_header_style(ws.cell(ws.max_row, i), gold=(i==1))
            for name, cost in by_item:
                ws.append([name, round(cost,2), f"{round(cost/total_cogs*100,1) if total_cogs else 0}%"])
            _xl_auto_width(ws)
            return StreamingResponse(_wb_to_stream(wb), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                     headers={"Content-Disposition": f"attachment; filename=food_cost_{ts}.xlsx"})
        else:
            summary = [["Period", f"Last {days} days"], ["Food Sales", f"${food_sales:,.2f}"], ["COGS", f"${total_cogs:,.2f}"],
                       ["Food Cost %", f"{food_pct}%"], ["Target", f"{target}%"], ["Variance", f"{round(food_pct-target,1)}%"]]
            data_rows = [[n, f"${c:,.2f}", f"{round(c/total_cogs*100,1) if total_cogs else 0}%"] for n, c in by_item]
            buf = _pdf_doc(f"Food Cost Report — Last {days} Days", data_rows, ["Item","Spend","% COGS"], summary_rows=summary)
            return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=food_cost_{ts}.pdf"})

    # ── WASTE ────────────────────────────────────────────────
    elif report == "waste":
        result  = await db.execute(select(WasteLog).where(WasteLog.date >= start).order_by(desc(WasteLog.date)))
        entries = result.scalars().all()
        total_cost = sum(e.estimated_cost or 0 for e in entries)

        if fmt == "xlsx":
            wb = openpyxl.Workbook()
            ws = wb.active; ws.title = "Waste Log"; ws.sheet_view.showGridLines = False
            _xl_title_row(ws, f"Waste Log — Last {days} Days", 6)
            ws.append([])
            headers = ["Date", "Item", "Type", "Reason", "Qty", "Est. Cost"]
            ws.append(headers)
            for i, h in enumerate(headers, 1): _xl_header_style(ws.cell(ws.max_row, i), gold=(i==1))
            for e in entries:
                ws.append([e.date.strftime("%Y-%m-%d"), e.item_name, e.item_type, e.reason, e.quantity, e.estimated_cost or 0])
            ws.append([]); ws.append(["", "", "", "", "TOTAL", round(total_cost,2)])
            c = ws.cell(ws.max_row, 6); c.font = Font(bold=True, color=_GOLD)
            _xl_auto_width(ws)
            return StreamingResponse(_wb_to_stream(wb), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                     headers={"Content-Disposition": f"attachment; filename=waste_{ts}.xlsx"})
        else:
            summary = [["Period", f"Last {days} days"], ["Total Entries", str(len(entries))], ["Total Waste Cost", f"${total_cost:,.2f}"]]
            data_rows = [[e.date.strftime("%Y-%m-%d"), e.item_name, e.item_type, e.reason, str(e.quantity or ""), f"${e.estimated_cost or 0:,.2f}"] for e in entries]
            buf = _pdf_doc(f"Waste Log — Last {days} Days", data_rows, ["Date","Item","Type","Reason","Qty","Cost"],
                           col_widths=[0.9*inch,1.6*inch,0.8*inch,0.9*inch,0.6*inch,0.8*inch], summary_rows=summary)
            return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=waste_{ts}.pdf"})

    # ── LOW STOCK ────────────────────────────────────────────
    elif report == "low-stock":
        # Reuse subquery pattern — no N+1
        bar_ts  = (select(InventoryCount.item_id, func.max(InventoryCount.timestamp).label("max_ts")).group_by(InventoryCount.item_id).subquery())
        bar_ct  = (select(InventoryCount.item_id, InventoryCount.level_percentage).join(bar_ts, and_(InventoryCount.item_id==bar_ts.c.item_id, InventoryCount.timestamp==bar_ts.c.max_ts)).subquery())
        bar_rows = (await db.execute(select(InventoryItem.name, InventoryItem.location, InventoryItem.category, bar_ct.c.level_percentage).outerjoin(bar_ct, InventoryItem.id==bar_ct.c.item_id).where(InventoryItem.is_active==True))).all()

        kit_ts  = (select(KitchenInventoryCount.item_id, func.max(KitchenInventoryCount.timestamp).label("max_ts")).group_by(KitchenInventoryCount.item_id).subquery())
        kit_ct  = (select(KitchenInventoryCount.item_id, KitchenInventoryCount.quantity).join(kit_ts, and_(KitchenInventoryCount.item_id==kit_ts.c.item_id, KitchenInventoryCount.timestamp==kit_ts.c.max_ts)).subquery())
        kit_rows = (await db.execute(select(KitchenInventoryItem.name, KitchenInventoryItem.location, KitchenInventoryItem.par_level, KitchenInventoryItem.unit, kit_ct.c.quantity).outerjoin(kit_ct, KitchenInventoryItem.id==kit_ct.c.item_id).where(KitchenInventoryItem.is_active==True))).all()

        low_bar = [(n, loc, cat, pct, "critical" if pct==0 else ("high" if pct<=10 else "medium"))
                   for n, loc, cat, pct in bar_rows if pct is not None and pct <= 25]
        low_kit = [(n, loc, f"{round(qty/par*100) if par else 0}% of par", "kitchen",
                    "critical" if qty==0 else ("high" if par and qty/par<=0.25 else "medium"))
                   for n, loc, par, unit, qty in kit_rows if qty is not None and par and qty < par]

        if fmt == "xlsx":
            wb = openpyxl.Workbook()
            ws = wb.active; ws.title = "Low Stock"; ws.sheet_view.showGridLines = False
            _xl_title_row(ws, "Low Stock Report", 5)
            ws.append([]); ws.append(["BAR ITEMS"])
            ws.cell(ws.max_row,1).font = Font(bold=True, color=_GOLD)
            headers = ["Item", "Location", "Category", "Level %", "Urgency"]
            ws.append(headers)
            for i, h in enumerate(headers, 1): _xl_header_style(ws.cell(ws.max_row, i), gold=(i==1))
            for n, loc, cat, pct, urg in low_bar:
                ws.append([n, loc or "", cat or "", f"{pct}%", urg.upper()])
            ws.append([]); ws.append(["KITCHEN ITEMS"])
            ws.cell(ws.max_row,1).font = Font(bold=True, color=_GOLD)
            headers2 = ["Item", "Location", "% of Par", "Type", "Urgency"]
            ws.append(headers2)
            for i, h in enumerate(headers2, 1): _xl_header_style(ws.cell(ws.max_row, i), gold=(i==1))
            for n, loc, pct_par, itype, urg in low_kit:
                ws.append([n, loc or "", pct_par, itype, urg.upper()])
            _xl_auto_width(ws)
            return StreamingResponse(_wb_to_stream(wb), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                     headers={"Content-Disposition": f"attachment; filename=low_stock_{ts}.xlsx"})
        else:
            summary = [["Bar Items Low", str(len(low_bar))], ["Kitchen Items Low", str(len(low_kit))], ["Total Low", str(len(low_bar)+len(low_kit))]]
            bar_data = [[n, loc or "", cat or "", f"{pct}%", urg.upper()] for n, loc, cat, pct, urg in low_bar]
            kit_data = [[n, loc or "", pct_par, itype, urg.upper()] for n, loc, pct_par, itype, urg in low_kit]
            buf = _pdf_doc("Low Stock Report", bar_data + kit_data, ["Item","Location","Detail","Type/Cat","Urgency"],
                           col_widths=[1.8*inch,1.2*inch,1.1*inch,1.0*inch,0.9*inch], summary_rows=summary)
            return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=low_stock_{ts}.pdf"})

    # ── VARIANCE ─────────────────────────────────────────────
    elif report == "variance":
        items_result = await db.execute(select(InventoryItem).where(InventoryItem.is_active == True))
        items = items_result.scalars().all()
        rows  = []
        for item in items:
            cr = await db.execute(
                select(InventoryCount.level_percentage, InventoryCount.timestamp)
                .where(InventoryCount.item_id == item.id).order_by(desc(InventoryCount.timestamp)).limit(2)
            )
            counts = cr.all()
            if len(counts) < 2: continue
            latest, previous = counts[0], counts[1]
            change = latest[0] - previous[0]
            cost_impact = round(abs(change)/100*(item.cost_per_unit or 0), 2)
            rows.append({"name": item.name, "location": item.location or "", "prev": previous[0], "curr": latest[0], "change": change, "cost": cost_impact, "flag": "shrinkage" if change < -25 else ("low" if latest[0] <= 25 else "ok")})
        rows.sort(key=lambda x: x["change"])

        if fmt == "xlsx":
            wb = openpyxl.Workbook()
            ws = wb.active; ws.title = "Variance"; ws.sheet_view.showGridLines = False
            _xl_title_row(ws, "Inventory Variance Report", 6)
            ws.append([])
            headers = ["Item", "Location", "Previous %", "Current %", "Change %", "Est. Cost Impact"]
            ws.append(headers)
            for i, h in enumerate(headers, 1): _xl_header_style(ws.cell(ws.max_row, i), gold=(i==1))
            for r in rows:
                ws.append([r["name"], r["location"], r["prev"], r["curr"], r["change"], r["cost"]])
            _xl_auto_width(ws)
            return StreamingResponse(_wb_to_stream(wb), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                     headers={"Content-Disposition": f"attachment; filename=variance_{ts}.xlsx"})
        else:
            summary = [["Items Tracked", str(len(rows))], ["Shrinkage Items", str(len([r for r in rows if r['flag']=='shrinkage']))],
                       ["Est. Shrinkage Cost", f"${sum(r['cost'] for r in rows if r['change']<-25):,.2f}"]]
            data_rows = [[r["name"], r["location"], f"{r['prev']}%", f"{r['curr']}%", f"{r['change']:+}%", f"${r['cost']:,.2f}"] for r in rows]
            buf = _pdf_doc("Inventory Variance Report", data_rows, ["Item","Location","Previous","Current","Change","Cost Impact"],
                           col_widths=[1.6*inch,1.0*inch,0.9*inch,0.9*inch,0.8*inch,1.0*inch], summary_rows=summary)
            return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=variance_{ts}.pdf"})

    else:
        raise HTTPException(status_code=400, detail=f"Unknown report type: {report}. Valid: sales, pour-cost, food-cost, waste, low-stock, variance")

# ============================================================
# Register router + CORS — MUST be after all @api_router decorators
# ============================================================
app.include_router(api_router)

_frontend_url = os.environ.get('FRONTEND_URL', '')
_allowed_origins = [o.strip() for o in _frontend_url.split(',') if o.strip()] if _frontend_url else []
if not _allowed_origins:
    logger.warning("FRONTEND_URL is not set — CORS will block all cross-origin requests.")

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=_allowed_origins,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["Authorization", "Content-Type", "Accept"],
)
